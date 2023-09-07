import openpyxl
import json
import re

def check_availability(value):
    # 如果value是None，返回"null"
    if value is None or value ==[]:
        return "null"
    # 如果value是字符串并且为空或只有空格，返回"null"
    elif isinstance(value, str) and value.strip() == "":
        return "null"
    # 对于其他情况，直接返回value
    else:
        return value


with open('input.json', 'r') as file:
    data = json.load(file)
    product_code_value = data.get('product_code')
   
    product_name = data.get('product_name')
    related_product_code = data.get('related_product_code')
    product_is_discontinued = data.get('product_is_discontinued')

    categories = data.get('categories')
    supplier_categories_category1 = categories.get('category1')
    supplier_categories_category2 = categories.get('category2')
    supplier_categories_category3 = categories.get('category3')
    supplier_categories_category4 = categories.get('category4')

    appa_category = data.get('appa_categories')

    split_parts = appa_category.split("/")
    first_part = split_parts[0]
    second_part = split_parts[1]
  
    short_description = data.get('short_description')
    full_description = data.get('full_description')


    tag = data.get('tag')

    tags = tag.split(",")

    promo = [t for t in tags if t in ["new", "sale", "trending"]]
    feature = [t for t in tags if t in ["eco", "full-colour"]]

    promo = ','.join(promo) if promo else 'null'
    feature = ','.join(feature) if feature else 'null'



    keywords = data.get('keywords')
    availbale_colour = data.get('availbale_colour')
    available_branding = data.get('available_branding')
    colour_pms= data.get('availbale_colour')

    specification_name1	= data.get('specification_name1')
    specification_value1 = data.get('specification_value1')
    specification_name2 = data.get('specification_name2')
    specification_value2 = data.get('specification_value2')
    specification_name3 = data.get('specification_name3')
    specification_value3 = data.get('specification_value3')
    specification_name4 = data.get('specification_name4')
    specification_value4 = data.get('specification_value4')
    specification_name5 = data.get('specification_name5')
    specification_value5 = data.get('specification_value5')



    packaging_type = data.get('packaging_type')
    carton_length = data.get('carton_length')
    carton_width = data.get('carton_width')
    carton_height = data.get('carton_height')
    carton_weight = data.get('carton_weight')
    carton_qty = data.get('carton_qty')



    images = data.get('images')

    filename_string = ""
    for image in images:
        url = image["url"]
        filename = url.split("/")[-1]
        filename_string += filename + ","

    if filename_string.endswith(","):
        image_string = filename_string[:-1]



    files = data.get('files')

    file_string = ""
    for file in files:
        url = file["url"]
        filename = url.split("/")[-1]
        file_string += filename + ","

    if filename_string.endswith(","):
        file_string = file_string[:-1]



    inventory_data = data.get('inventory')  # 获取 inventory 数据

    if isinstance(inventory_data, list):
        inventory_string = ""
        for item in inventory_data:
            item_name = item.get("itemName")
            if item_name:
                inventory_string += item_name + ","

        # 移除最后一个逗号
        if inventory_string.endswith(","):
            inventory_string = inventory_string[:-1]
    else:
        inventory_string = inventory_data.get("itemName")


    available_leadtime = data.get('carton_qty')


    leadtime_au_values = list(set(deco["leadtime_au"] for deco in data["decorations"]))

    # 定义一个函数来转换leadtime_au值
    def convert_leadtime(value):
        # 使用正则表达式匹配数字和后缀（如Hours或Days）
        match = re.search(r'(\d+)\s*(Hours|Working Days)', value)
        if match:
            number = match.group(1)
            suffix = 'H' if match.group(2) == 'Hours' else 'D'
            return f"EQ{number}{suffix}"
        return value  # 如果没有匹配，返回原值

    # 使用列表解析和convert_leadtime函数格式化每个值
    output= [convert_leadtime(val) for val in leadtime_au_values]
    available_leadtime = ', '.join(output)


    additional_info_price_disclaimer = data.get('price_disclaimer')
    freight_disclaimer_au = data.get('freight_disclaimer_au')
    freight_disclaimer_nz = data.get('freight_disclaimer_nz')
    additional_info = data.get('additional_info')
    change_log_au = data.get('change_log_au')
    change_log_nz = data.get('change_log_nz')


    combined_change_log = f"{change_log_au}{change_log_nz}"


    files = data.get('files')

    product_url = data.get('product_url')
    
    pricetable_au = data.get('pricetable_au')
    pricetable_nz = data.get('pricetable_nz')

    moq = pricetable_au[0]['moq']
    

    has_au = 'pricetable_au' in data
    has_nz = 'pricetable_nz' in data

    # 根据存在的键设置 available_leadtime 的值
    if has_au and has_nz:
        available_country = "AU, NZ"
    elif has_au:
        available_country = "AU"
    elif has_nz:
        available_country = "NZ"
    else:
        available_country = "null" 

    ava_au = 'Y' if 'AU' in available_country else 'N' if available_country else None
    ava_nz = 'Y' if 'NZ' in available_country else 'N' if available_country else None



supplier_code = data.get('supplier_code')


decorations = data.get('decorations')


video =  data.get('video')





# 加载现有的 Excel 文件
file_path = '/Users/wudongchen/Downloads/Open Promo - Data Design - V3 (1).xlsx'
workbook = openpyxl.load_workbook(file_path)

# 获取 "ProductsNew" 工作表
if "Products" in workbook.sheetnames:
    sheet = workbook["Products"]
else:
    print("Products sheet doesn't exist in the workbook.")
    exit()
suppier_name = "Dex Collection"
supplier_code = "DEX"

# 写入 "test" 到 "product_code" 列的下一行
# 假设 "product_code" 在第一行第一列，你需要根据实际情况调整
last_row = sheet.max_row


product_details = [
    suppier_name, supplier_code, product_name, product_code_value, related_product_code,
    product_is_discontinued, first_part,second_part, supplier_categories_category1, 
    supplier_categories_category2, supplier_categories_category3, 
    supplier_categories_category4,short_description, full_description,feature,promo,
    keywords, specification_name1, specification_value1, specification_name2,
    specification_value2, specification_name3, specification_value3, 
    specification_name4, specification_value4,specification_name5,specification_value5,availbale_colour, colour_pms,
    packaging_type, carton_length,carton_width, carton_height, carton_weight, carton_qty,moq,ava_au]

for col_num, value in enumerate(product_details, start=1):
    sheet.cell(row=last_row + 1, column=col_num).value = check_availability(value)

current_column = 38
for price in pricetable_au:
    description = check_availability(price.get('description', 'null'))
    moq_surcharge = check_availability(price.get('moq_surcharge', 'null'))
    
    qty_price_pairs = [
        (price.get(f'qty{i}', 'null'), price.get(f'price{i}', 'null'))
        for i in range(1, 10)
    ]
    
    # Update Excel cells
    sheet.cell(row=last_row + 1, column=current_column).value = description
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = moq_surcharge
    current_column += 1

    for qty, price in qty_price_pairs:
        sheet.cell(row=last_row + 1, column=current_column).value = qty
        current_column += 1
        
        sheet.cell(row=last_row + 1, column=current_column).value = price
        current_column += 1

sheet.cell(row=last_row + 1, column=58).value = ava_nz

current_column = 59
pricetable_nz = pricetable_nz or []
for price in pricetable_nz:
    description = check_availability(price.get('description', 'null'))
    moq_surcharge = check_availability(price.get('moq_surcharge', 'null'))
   
    qty_price_pairs = [
        (price.get(f'qty{i}', 'null'), price.get(f'price{i}', 'null'))
        for i in range(1, 10)
    ]
    
    # Update Excel cells
    sheet.cell(row=last_row + 1, column=current_column).value = description
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = moq_surcharge
    current_column += 1

    
    for qty, price in qty_price_pairs:
        sheet.cell(row=last_row + 1, column=current_column).value = qty
        current_column += 1
        
        sheet.cell(row=last_row + 1, column=current_column).value = price
        current_column += 1
    



current_column = 79
decorations = decorations or []
for dec in decorations:
    Name = check_availability(dec.get('Name', 'null'))
    Size = check_availability(dec.get('Size', 'null'))
   
    
    # Update Excel cells
    sheet.cell(row=last_row + 1, column=current_column).value = Name
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = Size
    current_column += 1

    

sheet.cell(row=last_row + 1, column=92).value = image_string
sheet.cell(row=last_row + 1, column=93).value = file_string  
sheet.cell(row=last_row + 1, column=94).value = video 
sheet.cell(row=last_row + 1, column=95).value = inventory_string
sheet.cell(row=last_row + 1, column=96).value = combined_change_log
sheet.cell(row=last_row + 1, column=97).value = additional_info




# 保存更改
workbook.save(file_path)
