import openpyxl
import json
import re

def check_availability(value):
    # 如果value是None，返回"null"
    if value is None:
        return "null"
    # 如果value是字符串并且为空或只有空格，返回"null"
    elif isinstance(value, str) and value.strip() == "":
        return "null"
    # 对于其他情况，直接返回value
    else:
        return value


with open('input.json', 'r') as file:
    data = json.load(file)
    product_code_value = data.get('product_code')
    product_name = data.get('product_name')
    related_product_code = data.get('related_product_code')
    product_is_discontinued = data.get('product_is_discontinued')

    categories = data.get('categories')
    supplier_categories_category1 = categories.get('category1')
    supplier_categories_category2 = categories.get('category2')
    supplier_categories_category3 = categories.get('category3')
    supplier_categories_category4 = categories.get('category4')
  
    short_description = data.get('short_description')
    full_description = data.get('full_description')
    tag = data.get('tag')
    keywords = data.get('keywords')
    availbale_colour = data.get('availbale_colour')
    available_branding = data.get('available_branding')
    colour_pms= data.get('availbale_colour')

    specification_name1	= data.get('specification_name1')
    specification_value1 = data.get('specification_value1')
    specification_name2 = data.get('specification_name2')
    specification_value2 = data.get('specification_value2')
    specification_name3 = data.get('specification_name3')
    specification_value3 = data.get('specification_value3')
    specification_name4 = data.get('specification_name4')
    specification_value4 = data.get('specification_value4')

    packaging_type = data.get('packaging_type')
    carton_length = data.get('carton_length')
    carton_width = data.get('carton_width')
    carton_height = data.get('carton_height')
    carton_weight = data.get('carton_weight')
    carton_qty = data.get('carton_qty')

    images = data.get('images')

    available_leadtime = data.get('carton_qty')


    leadtime_au_values = list(set(deco["leadtime_au"] for deco in data["decorations"]))

    # 定义一个函数来转换leadtime_au值
    def convert_leadtime(value):
        # 使用正则表达式匹配数字和后缀（如Hours或Days）
        match = re.search(r'(\d+)\s*(Hours|Working Days)', value)
        if match:
            number = match.group(1)
            suffix = 'H' if match.group(2) == 'Hours' else 'D'
            return f"EQ{number}{suffix}"
        return value  # 如果没有匹配，返回原值

    # 使用列表解析和convert_leadtime函数格式化每个值
    output= [convert_leadtime(val) for val in leadtime_au_values]
    available_leadtime = ', '.join(output)


    additional_info_price_disclaimer = data.get('price_disclaimer')
    freight_disclaimer_au = data.get('freight_disclaimer_au')
    freight_disclaimer_nz = data.get('freight_disclaimer_nz')
    additional_info = data.get('additional_info')
    change_log_au = data.get('change_log_au')
    change_log_nz = data.get('change_log_nz')

    files = data.get('files')

    product_url = data.get('product_url')
    
    pricetable_au = data.get('pricetable_au')
    pricetable_nz = data.get('pricetable_nz')
    

    has_au = 'pricetable_au' in data
    has_nz = 'pricetable_nz' in data

    # 根据存在的键设置 available_leadtime 的值
    if has_au and has_nz:
        available_country = "AU, NZ"
    elif has_au:
        available_country = "AU"
    elif has_nz:
        available_country = "NZ"
    else:
        available_country = "null" 

supplier_code = data.get('supplier_code')
# 加载现有的 Excel 文件
file_path = '/Users/wudongchen/Downloads/OpenPromo - Standard Data format - V1.xlsx'
workbook = openpyxl.load_workbook(file_path)

# 获取 "ProductsNew" 工作表
if "ProductsNew" in workbook.sheetnames:
    sheet = workbook["ProductsNew"]
else:
    print("ProductsNew sheet doesn't exist in the workbook.")
    exit()


# 写入 "test" 到 "product_code" 列的下一行
# 假设 "product_code" 在第一行第一列，你需要根据实际情况调整
last_row = sheet.max_row


product_details = [
    product_code_value, product_name, related_product_code,
    product_is_discontinued, supplier_categories_category1, 
    supplier_categories_category2, supplier_categories_category3, 
    supplier_categories_category4, short_description, full_description, tag,
    keywords, availbale_colour, available_branding, colour_pms, 
    specification_name1, specification_value1, specification_name2,
    specification_value2, specification_name3, specification_value3, 
    specification_name4, specification_value4, packaging_type, carton_length,
    carton_width, carton_height, carton_weight, carton_qty]

for col_num, value in enumerate(product_details, start=1):
    sheet.cell(row=last_row + 1, column=col_num).value = check_availability(value)


current_column = 30  # 假设在第30列开始写入URLs

for image in images:
    name = check_availability(image.get('name', 'null'))  # 使用null作为默认值
    tag = check_availability(image.get('tag', 'null'))
    colour = check_availability(image.get('colour', 'null'))
    url = check_availability(image.get('url', 'null'))
    
    sheet.cell(row=last_row + 1, column=current_column).value = name
    current_column += 1
    
    sheet.cell(row=last_row + 1, column=current_column).value = tag
    current_column += 1
    
    sheet.cell(row=last_row + 1, column=current_column).value = colour
    current_column += 1
    
    sheet.cell(row=last_row + 1, column=current_column).value = url
    current_column += 1

sheet.cell(row=last_row + 1, column=54).value = available_leadtime

sheet.cell(row=last_row + 1, column=55).value = additional_info_price_disclaimer
sheet.cell(row=last_row + 1, column=56).value = freight_disclaimer_au
sheet.cell(row=last_row + 1, column=57).value = freight_disclaimer_nz
sheet.cell(row=last_row + 1, column=58).value = additional_info
sheet.cell(row=last_row + 1, column=59).value = change_log_au
sheet.cell(row=last_row + 1, column=60).value = change_log_nz


current_column = 61  # 假设在第30列开始写入URLs

for file in files:
    name = check_availability(file.get('name', 'null')) # 使用null作为默认值
    tag = check_availability(file.get('tag', 'null'))
    url = check_availability(file.get('url', 'null'))
    
    sheet.cell(row=last_row + 1, column=current_column).value = name
    current_column += 1
    
    if name == 'ProductLineDrawing':
        sheet.cell(row=last_row + 1, column=current_column).value = "LineDrawing"
        current_column += 1
    elif name == 'ProductCertificate':
        sheet.cell(row=last_row + 1, column=current_column).value = "Certificate"
        current_column += 1
    else:
        # 如果tag不是LineDrawing或Certificate，存储默认值
        sheet.cell(row=last_row + 1, column=current_column).value = tag
        current_column += 1
    
    sheet.cell(row=last_row + 1, column=current_column).value = url
    current_column += 1

sheet.cell(row=last_row + 1, column=72).value = product_url
sheet.cell(row=last_row + 1, column=73).value = available_country
sheet.cell(row=last_row + 1, column=74).value = supplier_code


current_column = 75
for price in pricetable_au:
    description = check_availability(price.get('description', 'null'))
    instruction = check_availability(price.get('instruction', 'null'))
    moq = check_availability(price.get('moq', 'null'))
    moq_surcharge = check_availability(price.get('moq_surcharge', 'null'))
    lowest_price = check_availability(price.get('lowest_price', 'null'))
    
    qty_price_pairs = [
        (price.get(f'qty{i}', 'null'), price.get(f'price{i}', 'null'))
        for i in range(1, 10)
    ]
    
    # Update Excel cells
    sheet.cell(row=last_row + 1, column=current_column).value = description
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = instruction
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = moq
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = moq_surcharge
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = lowest_price
    current_column += 1
    
    for qty, price in qty_price_pairs:
        sheet.cell(row=last_row + 1, column=current_column).value = qty
        current_column += 1
        
        sheet.cell(row=last_row + 1, column=current_column).value = price
        current_column += 1
    
    # If you want to reset the current_column to a specific number after each loop, do it here.
    # Otherwise, remove or comment the next line.
    # current_column = INITIAL_COLUMN_VALUE


current_column = 121
pricetable_nz = pricetable_nz or []
for price in pricetable_nz:
    description = check_availability(price.get('description', 'null'))
    instruction = check_availability(price.get('instruction', 'null'))
    moq = check_availability(price.get('moq', 'null'))
    moq_surcharge = check_availability(price.get('moq_surcharge', 'null'))
    lowest_price = check_availability(price.get('lowest_price', 'null'))
    
    qty_price_pairs = [
        (price.get(f'qty{i}', 'null'), price.get(f'price{i}', 'null'))
        for i in range(1, 10)
    ]
    
    # Update Excel cells
    sheet.cell(row=last_row + 1, column=current_column).value = description
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = instruction
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = moq
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = moq_surcharge
    current_column += 1

    sheet.cell(row=last_row + 1, column=current_column).value = lowest_price
    current_column += 1
    
    for qty, price in qty_price_pairs:
        sheet.cell(row=last_row + 1, column=current_column).value = qty
        current_column += 1
        
        sheet.cell(row=last_row + 1, column=current_column).value = price
        current_column += 1
    
    
# 保存更改
workbook.save(file_path)
